"""
Excel处理模块 - 负责题目的导入和导出
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Optional
import os


def detect_question_type(question: str, option_c: str, option_d: str) -> str:
    """
    根据题目内容自动检测题型
    
    Args:
        question: 题干
        option_c: 选项C
        option_d: 选项D
        
    Returns:
        题型: single/multiple/judge
    """
    # 判断题特征：题干包含"是否正确"等字样，或只有A/B两个选项且内容包含"正确/错误"
    question_lower = question.lower()
    if '是否正确' in question or '以下说法' in question:
        if not option_c and not option_d:
            return 'judge'
    
    # 检查选项内容是否为判断类型
    if option_c == '' and option_d == '':
        return 'judge'
    
    return 'single'


def import_from_excel(filepath: str) -> List[Dict]:
    """
    从Excel文件导入题目
    
    列格式：
    A: 序号
    B: 题型 (single/multiple/judge，可选)
    C: 题干
    D-I: 选项A-F
    J: 答案
    
    Args:
        filepath: Excel文件路径
        
    Returns:
        题目列表
        
    Raises:
        ValueError: 数据格式错误，包含具体行号和错误信息
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"文件不存在: {filepath}")
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    questions = []
    row_num = 1  # 用于错误提示
    
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        row_num += 1
        
        if len(row) < 10 or row[0] is None:
            continue
        
        seq_no = int(row[0]) if isinstance(row[0], (int, float)) else None
        if seq_no is None:
            continue
        
        # 读取题型，如果没有则自动检测
        q_type = row[1] if row[1] else None
        
        question_text = str(row[2]) if row[2] else ""
        
        # 读取选项
        options = []
        for i in range(3, 9):  # D-I列，索引3-8
            opt = str(row[i]) if row[i] and str(row[i]).strip() else ""
            options.append(opt)
        
        # 读取答案
        answer = str(row[9]).upper().strip() if row[9] else ""
        
        # 如果未指定题型，自动检测
        if not q_type:
            q_type = detect_question_type(question_text, options[2], options[3])
        else:
            q_type = str(q_type).lower().strip()
            if q_type in ['单选', '单选题', 'single']:
                q_type = 'single'
            elif q_type in ['多选', '多选题', 'multiple']:
                q_type = 'multiple'
            elif q_type in ['判断', '判断题', 'judge']:
                q_type = 'judge'
            else:
                # 无法识别的题型，尝试自动检测
                q_type = detect_question_type(question_text, options[2], options[3])
        
        # 确保题型不为空
        if not q_type:
            raise ValueError(f"第 {row_num} 行：无法识别题型，请检查B列是否填写正确（single/multiple/judge 或 单选/多选/判断）")
        
        # 确保题干不为空
        if not question_text.strip():
            raise ValueError(f"第 {row_num} 行：题干（C列）不能为空")
        
        # 确保答案不为空
        if not answer:
            raise ValueError(f"第 {row_num} 行：答案（J列）不能为空")
        
        # 如果是判断题且没有选项，添加默认选项
        if q_type == 'judge':
            if not options[0]:
                options[0] = "正确"
            if not options[1]:
                options[1] = "错误"
        
        # 根据答案长度判断是否为多选题
        if len(answer) > 1 and q_type == 'single':
            q_type = 'multiple'
        
        q = {
            'seq_no': seq_no,
            'type': q_type,
            'question': question_text,
            'option_a': options[0],
            'option_b': options[1],
            'option_c': options[2],
            'option_d': options[3],
            'option_e': options[4],
            'option_f': options[5],
            'answer': answer
        }
        
        questions.append(q)
    
    wb.close()
    return questions


def export_to_excel(filepath: str, questions: List[Dict], 
                    include_wrong_count: bool = True,
                    include_time_spent: bool = True):
    """
    导出题目到Excel
    
    列格式：
    A: 序号
    B: 题型
    C: 题干
    D-I: 选项A-F
    J: 答案
    K: 错题次数 (可选)
    L: 做题时间秒数 (可选)
    
    Args:
        filepath: 导出文件路径
        questions: 题目列表
        include_wrong_count: 是否包含错题次数
        include_time_spent: 是否包含做题时间
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题目"
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头
    headers = ["序号", "题型", "题干", "选项A", "选项B", "选项C", 
               "选项D", "选项E", "选项F", "答案"]
    
    if include_wrong_count:
        headers.append("错题次数")
    if include_time_spent:
        headers.append("做题时间秒数")
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # 写入数据
    type_names = {
        'single': '单选题',
        'multiple': '多选题', 
        'judge': '判断题'
    }
    
    for row_idx, q in enumerate(questions, 2):
        ws.cell(row=row_idx, column=1, value=q.get('seq_no'))
        ws.cell(row=row_idx, column=2, value=type_names.get(q.get('type'), q.get('type')))
        ws.cell(row=row_idx, column=3, value=q.get('question'))
        ws.cell(row=row_idx, column=4, value=q.get('option_a', ''))
        ws.cell(row=row_idx, column=5, value=q.get('option_b', ''))
        ws.cell(row=row_idx, column=6, value=q.get('option_c', ''))
        ws.cell(row=row_idx, column=7, value=q.get('option_d', ''))
        ws.cell(row=row_idx, column=8, value=q.get('option_e', ''))
        ws.cell(row=row_idx, column=9, value=q.get('option_f', ''))
        ws.cell(row=row_idx, column=10, value=q.get('answer', ''))
        
        col_idx = 11
        if include_wrong_count:
            ws.cell(row=row_idx, column=col_idx, value=q.get('wrong_count', 0))
            col_idx += 1
        if include_time_spent:
            time_val = q.get('time_spent', 0)
            ws.cell(row=row_idx, column=col_idx, value=round(time_val, 1) if time_val else 0)
        
        # 应用边框和对齐
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = thin_border
            if c in [1, 2, 10, 11, 12]:  # 数字和短文本居中
                cell.alignment = center_align
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 调整列宽
    col_widths = [8, 10, 40, 20, 20, 20, 20, 20, 20, 10, 12, 15]
    for i, width in enumerate(col_widths[:len(headers)], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    for row in range(2, len(questions) + 2):
        ws.row_dimensions[row].height = 30
    
    wb.save(filepath)
    wb.close()


def create_sample_excel(filepath: str):
    """
    创建示例Excel文件
    
    Args:
        filepath: 文件保存路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题目"
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    
    # 表头
    headers = ["序号", "题型", "题干", "选项A", "选项B", "选项C", 
               "选项D", "选项E", "选项F", "答案"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # 示例数据
    sample_data = [
        [1, "single", "Python中，以下哪个函数用于获取列表的长度？",
         "len()", "length()", "size()", "count()", "", "", "A"],
        [2, "single", "在HTML中，用于定义最大标题的标签是？",
         "\u003chead\u003e", "\u003cheader\u003e", "\u003ch1\u003e", "\u003ctitle\u003e", "", "", "C"],
        [3, "multiple", "以下哪些是Python的数据类型？",
         "int", "str", "array", "list", "class", "", "ABD"],
        [4, "judge", "CSS是用于网页结构设计的语言。",
         "正确", "错误", "", "", "", "", "B"],
        [5, "multiple", "以下哪些是前端开发框架？",
         "React", "Vue", "Django", "Angular", "Flask", "Spring", "ABD"],
        [6, "single", "HTTP状态码404表示？",
         "服务器错误", "请求成功", "未找到资源", "权限不足", "", "", "C"],
        [7, "judge", "JavaScript是一种编译型语言。",
         "正确", "错误", "", "", "", "", "B"],
        [8, "single", "在Git中，用于将更改提交到本地仓库的命令是？",
         "git push", "git commit", "git add", "git pull", "", "", "B"],
        [9, "multiple", "以下哪些是数据库管理系统？",
         "MySQL", "MongoDB", "Redis", "Excel", "Oracle", "", "ABCE"],
        [10, "single", "在Python中，以下哪个关键字用于定义函数？",
         "function", "def", "func", "define", "", "", "B"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 调整列宽
    col_widths = [8, 10, 45, 18, 18, 18, 18, 18, 18, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    wb.save(filepath)
    wb.close()
